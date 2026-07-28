from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any

from app.data_model import RegionalStatisticData

from .parser import parse_official_text


HEADER_ALIASES = {
    "指标编码": "metric_code",
    "指标名称": "metric_name",
    "数值": "value_numeric",
    "文字值": "value_text",
    "单位": "unit",
    "范围层级": "scope_level",
    "行政区代码": "scope_code",
    "行政区名称": "scope_name",
    "统计期": "stat_period",
}

REQUIRED_HEADERS = {
    "metric_code",
    "metric_name",
    "scope_level",
    "scope_code",
    "scope_name",
    "stat_period",
}


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        normalized[HEADER_ALIASES.get(str(key).strip(), str(key).strip())] = value
    return normalized


def _row_to_item(
    row: dict[str, Any],
    *,
    source_name: str,
    source_url: str,
    source_format: str,
) -> RegionalStatisticData:
    row = _normalize_row(row)
    missing = [key for key in REQUIRED_HEADERS if not str(row.get(key) or "").strip()]
    if missing:
        raise ValueError(f"缺少字段：{', '.join(sorted(missing))}")
    numeric = row.get("value_numeric")
    value_numeric = None if numeric in (None, "") else float(str(numeric).replace(",", ""))
    return RegionalStatisticData(
        metric_code=str(row["metric_code"]).strip(),
        metric_name=str(row["metric_name"]).strip(),
        value_numeric=value_numeric,
        value_text=str(row.get("value_text") or "").strip() or None,
        unit=str(row.get("unit") or "").strip() or None,
        scope_level=str(row["scope_level"]).strip(),
        scope_code=str(row["scope_code"]).strip(),
        scope_name=str(row["scope_name"]).strip(),
        stat_period=str(row["stat_period"]).strip(),
        source_name=source_name,
        source_url=source_url,
        source_format=source_format,
        source="government_stats",
        status="confirmed",
        confidence=.95 if source_format in {"csv", "xlsx"} else .8,
        timestamp=datetime.now(timezone.utc),
        raw_data={"uploaded": True},
    )


def parse_csv_upload(
    content: bytes,
    *,
    source_name: str,
    source_url: str,
) -> tuple[list[RegionalStatisticData], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    items: list[RegionalStatisticData] = []
    errors: list[dict[str, Any]] = []
    for row_number, row in enumerate(reader, start=2):
        try:
            items.append(
                _row_to_item(
                    row,
                    source_name=source_name,
                    source_url=source_url,
                    source_format="csv",
                )
            )
        except (TypeError, ValueError) as exc:
            errors.append({"row": row_number, "reason": str(exc)})
    return items, errors


def parse_xlsx_upload(
    content: bytes,
    *,
    source_name: str,
    source_url: str,
) -> tuple[list[RegionalStatisticData], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX 解析依赖 openpyxl 未安装，请先安装 backend requirements.txt") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows, ())]
    items: list[RegionalStatisticData] = []
    errors: list[dict[str, Any]] = []
    for row_number, values in enumerate(rows, start=2):
        row = dict(zip(headers, values))
        if not any(value not in (None, "") for value in values):
            continue
        try:
            items.append(
                _row_to_item(
                    row,
                    source_name=source_name,
                    source_url=source_url,
                    source_format="xlsx",
                )
            )
        except (TypeError, ValueError) as exc:
            errors.append({"row": row_number, "reason": str(exc)})
    return items, errors


def parse_pdf_upload(
    content: bytes,
    *,
    source_name: str,
    source_url: str,
    scope_level: str,
    scope_code: str,
    scope_name: str,
    stat_period: str,
) -> tuple[list[RegionalStatisticData], list[dict[str, Any]]]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("PDF 解析依赖 pypdf 未安装，请先安装 backend requirements.txt") from exc

    reader = PdfReader(io.BytesIO(content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    items = parse_official_text(
        text,
        scope_level=scope_level,
        scope_code=scope_code,
        scope_name=scope_name,
        source_name=source_name,
        source_url=source_url,
        source_format="pdf",
        stat_period=stat_period,
    )
    errors = [] if items else [{"row": 0, "reason": "PDF未识别到目标指标，请检查是否为扫描图片或调整文件格式"}]
    return items, errors


class GovernmentUploadAdapter:
    """管理员上传官方文件的确定性解析入口。"""

    def parse(
        self,
        filename: str,
        content: bytes,
        *,
        source_name: str,
        source_url: str,
        scope_level: str = "city",
        scope_code: str = "610100",
        scope_name: str = "西安市",
        stat_period: str = "",
    ) -> tuple[list[RegionalStatisticData], list[dict[str, Any]]]:
        suffix = filename.lower()
        if suffix.endswith(".csv"):
            return parse_csv_upload(content, source_name=source_name, source_url=source_url)
        if suffix.endswith(".xlsx"):
            return parse_xlsx_upload(content, source_name=source_name, source_url=source_url)
        if suffix.endswith(".pdf"):
            return parse_pdf_upload(
                content,
                source_name=source_name,
                source_url=source_url,
                scope_level=scope_level,
                scope_code=scope_code,
                scope_name=scope_name,
                stat_period=stat_period,
            )
        raise ValueError("仅支持 CSV、XLSX 或 PDF 文件")
